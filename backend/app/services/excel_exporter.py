from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


TITLE_FONT = Font(bold=True, size=15, color="111827")
SUBTITLE_FONT = Font(italic=True, size=10, color="6B7280")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SECTION_FONT = Font(bold=True, size=11, color="111827")

TITLE_FILL = PatternFill(fill_type="solid", fgColor="E2E8F0")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="1E293B")
ALT_ROW_FILL = PatternFill(fill_type="solid", fgColor="F8FAFC")
TOTAL_FILL = PatternFill(fill_type="solid", fgColor="FEF3C7")

BODY_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGNMENT = Alignment(horizontal="left", vertical="top", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)


def _safe_sheet_name(name: str) -> str:
    bad = ['\\', '/', '*', '?', ':', '[', ']']
    out = str(name or "Sheet")
    for ch in bad:
        out = out.replace(ch, " ")
    out = " ".join(out.split()).strip() or "Sheet"
    return out[:31]


def _normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\r", " ").replace("\n", " ").strip()


def _looks_like_total_label(text: str) -> bool:
    text = (text or "").strip().lower()
    return any(
        key in text
        for key in [
            "total",
            "grand total",
            "subtotal",
            "balance",
            "closing balance",
            "opening balance",
            "amount due",
            "net amount",
            "tax total",
        ]
    )


def _looks_numeric(value: Any) -> bool:
    text = _normalize_text(value)
    if not text:
        return False

    cleaned = (
        text.replace(",", "")
        .replace("₹", "")
        .replace("$", "")
        .replace("€", "")
        .replace("£", "")
        .replace("%", "")
        .replace("(", "-")
        .replace(")", "")
        .strip()
    )

    if cleaned.count(".") > 1:
        return False

    try:
        float(cleaned)
        return True
    except Exception:
        return False


def _to_number(value: Any) -> float | None:
    text = _normalize_text(value)
    if not text:
        return None

    cleaned = (
        text.replace(",", "")
        .replace("₹", "")
        .replace("$", "")
        .replace("€", "")
        .replace("£", "")
        .replace("%", "")
        .replace("(", "-")
        .replace(")", "")
        .strip()
    )

    try:
        return float(cleaned)
    except Exception:
        return None


def _is_mostly_numeric_row(row: list[Any]) -> bool:
    non_empty = [cell for cell in row if _normalize_text(cell)]
    if not non_empty:
        return False

    numeric_count = sum(1 for cell in non_empty if _looks_numeric(cell))
    return numeric_count >= max(1, len(non_empty) // 2)


def _pad_rows(rows: list[list[Any]], total_cols: int) -> list[list[str]]:
    padded: list[list[str]] = []

    for row in rows:
        normalized = [_normalize_text(v) for v in row]
        if len(normalized) < total_cols:
            normalized.extend([""] * (total_cols - len(normalized)))
        elif len(normalized) > total_cols:
            normalized = normalized[:total_cols]
        padded.append(normalized)

    return padded


def _autosize(ws) -> None:
    widths: dict[int, int] = {}

    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue

            value = str(cell.value)
            col_idx = cell.column
            current = widths.get(col_idx, 0)

            max_line = max(len(line) for line in value.splitlines()) if value else 0
            widths[col_idx] = max(current, max_line)

    for col_idx, max_len in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 40)


def _style_cell(cell, *, header: bool = False, alternate: bool = False, total: bool = False) -> None:
    cell.border = THIN_BORDER

    if header:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGNMENT
        return

    if total:
        cell.fill = TOTAL_FILL
        cell.font = Font(bold=True, color="111827")
        cell.alignment = BODY_ALIGNMENT
        return

    if alternate:
        cell.fill = ALT_ROW_FILL

    cell.alignment = BODY_ALIGNMENT


def _write_title_block(ws, workbook_title: str, sheet_name: str, kind: str | None) -> None:
    ws["A1"] = workbook_title
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = TITLE_FILL

    ws["A2"] = f"Sheet: {sheet_name}"
    ws["A2"].font = SECTION_FONT

    meta_text = f"Type: {kind}" if kind else "Type: extracted"
    ws["A3"] = meta_text
    ws["A3"].font = SUBTITLE_FONT


def _write_empty_sheet(ws) -> None:
    ws["A5"] = "No rows available"
    ws["A5"].font = Font(bold=True, color="6B7280")


def _apply_numeric_format(ws, start_row: int, end_row: int, total_cols: int) -> None:
    for row_idx in range(start_row, end_row + 1):
        for col_idx in range(1, total_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'


def _convert_numeric_cells(rows: list[list[str]]) -> list[list[Any]]:
    converted: list[list[Any]] = []

    for row in rows:
        out_row: list[Any] = []
        for value in row:
            num = _to_number(value)
            if num is not None and value.strip():
                out_row.append(num)
            else:
                out_row.append(value)
        converted.append(out_row)

    return converted


def export_workbook_plan_to_excel(
    workbook_plan: dict[str, Any],
    output_path: str | Path,
) -> str:
    output_path = str(output_path)
    wb = Workbook()

    default_ws = wb.active
    wb.remove(default_ws)

    workbook_title = _normalize_text(workbook_plan.get("workbook_title") or "Extracted Data")
    sheets = workbook_plan.get("sheets", []) or []

    if not sheets:
        ws = wb.create_sheet("Summary")
        _write_title_block(ws, workbook_title, "Summary", "summary")
        ws["A5"] = "No extractable data found"
        ws["A5"].font = Font(bold=True, color="6B7280")
        ws.freeze_panes = "A6"
        ws.sheet_view.showGridLines = True
        _autosize(ws)
        wb.save(output_path)
        return output_path

    used_names: set[str] = set()

    for idx, sheet in enumerate(sheets, start=1):
        raw_name = sheet.get("name", f"Sheet {idx}")
        name = _safe_sheet_name(raw_name)

        original_name = name
        suffix = 2
        while name in used_names:
            trimmed = original_name[: max(1, 31 - len(str(suffix)) - 1)]
            name = f"{trimmed} {suffix}"[:31]
            suffix += 1
        used_names.add(name)

        kind = _normalize_text(sheet.get("kind") or "table")
        columns = [_normalize_text(col) for col in (sheet.get("columns", []) or [])]
        rows = sheet.get("rows", []) or []

        total_cols = max(len(columns), max((len(r) for r in rows), default=0), 1)
        if not columns:
            columns = [f"Column {i}" for i in range(1, total_cols + 1)]

        normalized_rows = _pad_rows(rows, total_cols)
        converted_rows = _convert_numeric_cells(normalized_rows)

        ws = wb.create_sheet(name)
        _write_title_block(ws, workbook_title, name, kind)

        header_row = 5
        data_start_row = 6

        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=col_name)
            _style_cell(cell, header=True)

        if not converted_rows:
            _write_empty_sheet(ws)
            ws.freeze_panes = "A6"
            ws.sheet_view.showGridLines = True
            _autosize(ws)
            continue

        for row_offset, row in enumerate(converted_rows, start=0):
            excel_row = data_start_row + row_offset

            row_text_values = [_normalize_text(v) for v in row]
            first_text = row_text_values[0] if row_text_values else ""
            is_total_row = _looks_like_total_label(first_text)
            alternate = row_offset % 2 == 1

            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=excel_row, column=col_idx, value=value)
                _style_cell(cell, alternate=alternate, total=is_total_row)

                if col_idx == 1 and is_total_row:
                    cell.font = Font(bold=True, color="111827")

            if _is_mostly_numeric_row(row_text_values) and not is_total_row:
                for col_idx in range(1, total_cols + 1):
                    ws.cell(row=excel_row, column=col_idx).alignment = BODY_ALIGNMENT

        end_row = data_start_row + len(converted_rows) - 1
        _apply_numeric_format(ws, data_start_row, end_row, total_cols)

        ws.freeze_panes = "A6"
        ws.sheet_view.showGridLines = True
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(total_cols)}{end_row}"

        for col_idx in range(1, total_cols + 1):
            header_cell = ws.cell(row=header_row, column=col_idx)
            body_cells = [ws.cell(row=r, column=col_idx) for r in range(data_start_row, end_row + 1)]

            max_len = len(str(header_cell.value or ""))
            for cell in body_cells:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))

            width = min(max(max_len + 2, 12), 40)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.row_dimensions[1].height = 22
        ws.row_dimensions[2].height = 20
        ws.row_dimensions[3].height = 18
        ws.row_dimensions[5].height = 22

    wb.save(output_path)
    return output_path